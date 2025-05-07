import os
import zipfile
import platform
import json
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from tkinter import font as tkfont
from openpyxl import load_workbook

recent_folders = []
favorite_folders = []

if platform.system() == "Windows":
    CONFIG_DIR = os.path.join(os.getenv("APPDATA"), "ExcelSearcher")
else:
    CONFIG_DIR = os.path.join(os.path.expanduser("~/.config"), "excelsearcher")
os.makedirs(CONFIG_DIR, exist_ok=True)
CONFIG_PATH = os.path.join(CONFIG_DIR, "config.json")


def load_config():
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
            return config
    except:
        return {}

def save_config():
    config = {
        "font_size": tree_font_size,
        "recent_folders": recent_folders,
        "favorite_folders": favorite_folders
    }
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f)



def select_folder():
    folder = filedialog.askdirectory()
    if folder:
        folder_var.set(folder)
        update_recent_folders(folder)
        update_fav_button()


def update_recent_folders(folder):
    if folder not in recent_folders:
        recent_folders.insert(0, folder)
        if len(recent_folders) > 5:
            recent_folders.pop()
        folder_dropdown["values"] = recent_folders
        save_config()

def set_folder_from_dropdown(event):
    folder_var.set(folder_dropdown.get())
    update_fav_button()

def clear_treeview():
    for item in result_tree.get_children():
        result_tree.delete(item)

def search_excel():
    folder_path = folder_var.get()
    search_term = search_var.get().strip()

    if not folder_path or not search_term:
        status_label.config(text="Please select a folder and enter a search term.")
        return
    else:
        status_label.config(text="")  # Clear previous messages


    clear_treeview()

    # Preprocess search terms

    if match_any_var.get():
        search_terms = search_term.split()
    else:
        search_terms = [search_term]

    def normalize(s):
        s = str(s)
        return s if case_sensitive_var.get() else s.lower()

    search_terms = [normalize(term) for term in search_terms]


    for filename in os.listdir(folder_path):
        if filename.startswith("~$"):
            continue  # Skip temp Excel lock files
        if filename.endswith(".xlsx"):
            filepath = os.path.join(folder_path, filename)

            if not zipfile.is_zipfile(filepath):
                print(f"Skipping non-valid Excel file: {filename}")
                continue

            try:
                wb = load_workbook(filepath, read_only=True)
                file_has_match = False
                file_node = None

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        if not row:
                            continue

                        # Check for match
                        if any(
                            any(term in normalize(cell) for term in search_terms)
                            for cell in row if cell
                        ):
                            row_data = [str(cell).replace("\n", " ").replace("\r", " ") if cell else "" for cell in row[:10]]
                            if not file_has_match:
                                file_node = result_tree.insert("", "end", text=f"📁 {filename}", open=True, values=[""] * len(column_names))
                                file_has_match = True
                            result_tree.insert(file_node, "end", text=f"[{sheet_name}]", values=row_data)
                wb.close()
                auto_resize_columns()
            except Exception as e:
                print(f"Error reading {filename}: {e}")

def auto_resize_columns():
    padding = 20
    max_widths = [0] * len(column_names)
    font_obj = tree_font

    def measure_row(item_id):
        values = result_tree.item(item_id, "values")
        if not values: return
        for i, value in enumerate(values):
            text = str(value)
            width = font_obj.measure(text)
            max_widths[i] = max(max_widths[i], width)
        for child_id in result_tree.get_children(item_id):
            measure_row(child_id)

    for parent_id in result_tree.get_children():
        measure_row(parent_id)

    for i, col in enumerate(column_names):
        width = max(60, min(max_widths[i] + padding, 300))  # clamp between 60 and 300
        result_tree.column(col, width=width)





def enable_drag_scroll(widget):
    def scroll_start(event):
        widget.scan_mark(event.x, event.y)
    def scroll_move(event):
        widget.scan_dragto(event.x, event.y, gain=1)

    widget.bind("<ButtonPress-2>", scroll_start)  # Middle-click
    widget.bind("<B2-Motion>", scroll_move)

def open_excel_file(filepath):
    try:
        if platform.system() == "Windows":
            os.startfile(filepath)
        elif platform.system() == "Darwin":  # macOS
            import subprocess
            subprocess.call(["open", filepath])
        else:  # Linux
            import subprocess
            subprocess.call(["xdg-open", filepath])
    except Exception as e:
        print(f"Could not open file: {e}")

def on_row_double_click(event):
    selected = result_tree.selection()
    if not selected:
        return

    item_id = selected[0]
    parent_id = result_tree.parent(item_id)
    if not parent_id:
        return  # clicked on file-level row, not a result

    # Get file path from parent node
    file_label = result_tree.item(parent_id)["text"]
    file_path = os.path.join(folder_var.get(), file_label.replace("📁 ", ""))
    open_excel_file(file_path)


# --- GUI SETUP ---
root = tk.Tk()
root.title("Excel Text Searcher")
root.geometry("1100x650")

# Set font for proper multilingual support
try:
    root.option_add("*Font", ("Malgun Gothic", 10))  # Windows
except:
    root.option_add("*Font", ("Arial Unicode MS", 10))  # Mac fallback

folder_var = tk.StringVar()
search_var = tk.StringVar()
match_any_var = tk.BooleanVar(value=False)
case_sensitive_var = tk.BooleanVar(value=False)


# Folder & recent dropdown
frame_top = tk.Frame(root)
frame_top.pack(fill="x", padx=10, pady=5)

tk.Label(frame_top, text="Folder:").pack(side="left")
folder_dropdown = ttk.Combobox(frame_top, textvariable=folder_var, width=60, postcommand=lambda: folder_dropdown.set(folder_var.get()))
folder_dropdown.pack(side="left", padx=5)
folder_dropdown.bind("<<ComboboxSelected>>", set_folder_from_dropdown)
tk.Button(frame_top, text="Browse", command=select_folder).pack(side="left", padx=(5, 2))

def toggle_favorite():
    folder = folder_var.get()
    if not folder:
        return
    if folder in favorite_folders:
        favorite_folders.remove(folder)
    else:
        favorite_folders.insert(0, folder)
    save_config()
    update_fav_button()
    refresh_favorite_menu()

def update_fav_button():
    folder = folder_var.get()
    if folder in favorite_folders:
        fav_button.config(text="★")  # filled star
    else:
        fav_button.config(text="☆")  # empty star

fav_button = tk.Button(frame_top, text="☆", command=toggle_favorite)
fav_button.pack(side="left")

fav_menu_button = tk.Menubutton(frame_top, text="▼", relief="raised")
fav_menu = tk.Menu(fav_menu_button, tearoff=0)
fav_menu_button["menu"] = fav_menu
fav_menu_button.pack(side="left", padx=(2, 0))

def refresh_favorite_menu():
    fav_menu.delete(0, "end")
    if not favorite_folders:
        fav_menu.add_command(label="(No favorites)", state="disabled")
        return
    for folder in favorite_folders:
        label = os.path.basename(folder) or folder
        fav_menu.add_command(
            label=label,
            command=lambda f=folder: select_favorite_folder(f)
        )

def select_favorite_folder(folder):
    folder_var.set(folder)
    update_recent_folders(folder)
    update_fav_button()


# Search input
frame_search = tk.Frame(root)
frame_search.pack(fill="x", padx=10, pady=5)

tk.Label(frame_search, text="Search Term:").pack(side="left")
search_entry = tk.Entry(frame_search, textvariable=search_var, width=40)
search_entry.pack(side="left", padx=5)
search_entry.bind("<Return>", lambda event: search_excel())
search_button = tk.Button(frame_search, text="Search", command=search_excel, bg="#4CAF50", )
search_button.pack(side="left", padx=(5, 10))

status_label = tk.Label(frame_search, text="", fg="red")
status_label.pack(side="left")

options_frame = tk.Frame(root)
options_frame.pack(anchor="w", padx=10)

tk.Checkbutton(options_frame, text="Any of words", variable=match_any_var).pack(side="left", padx=5)
tk.Checkbutton(options_frame, text="Case sensitive", variable=case_sensitive_var).pack(side="left", padx=5)


# Zoom buttons below search bar
zoom_frame = tk.Frame(root)
zoom_frame.pack(anchor="w", padx=10, pady=(0, 5))

def zoom_in():
    global tree_font_size
    tree_font_size += 1
    tree_font.configure(size=tree_font_size)
    style.configure("Treeview", font=tree_font, rowheight=tree_font_size + 8)
    style.configure("Treeview.Heading", font=(tree_font_name, tree_font_size, "bold"))
    save_config()

def zoom_out():
    global tree_font_size
    if tree_font_size > 6:
        tree_font_size -= 1
        tree_font.configure(size=tree_font_size)
        style.configure("Treeview", font=tree_font, rowheight=tree_font_size + 8)
        style.configure("Treeview.Heading", font=(tree_font_name, tree_font_size, "bold"))
        save_config()


tk.Button(zoom_frame, text="Zoom In", command=zoom_in).pack(side="left", padx=2)
tk.Button(zoom_frame, text="Zoom Out", command=zoom_out).pack(side="left", padx=2)


# Define Treeview columns
column_names = [f"Col {i+1}" for i in range(10)]
# Treeview inside a frame with proper grid layout
tree_frame = tk.Frame(root)
tree_frame.pack(fill="both", expand=True, padx=10, pady=(5, 0))

result_tree = ttk.Treeview(tree_frame, columns=column_names, show="tree headings")
result_tree["displaycolumns"] = column_names
result_tree.bind("<Double-1>", on_row_double_click)

# Prevent first column from collapsing
result_tree.column("#0", width=140, minwidth=100, stretch=False)
result_tree.heading("#0", text="File / Sheet")


def refresh_scrollbars(event=None):
    result_tree.update_idletasks()

result_tree.bind("<Configure>", refresh_scrollbars)
result_tree.bind("<ButtonRelease-1>", refresh_scrollbars)  # catches manual column drag



from tkinter import font as tkfont

# Set initial font size and font name
config = load_config()
recent_folders = config.get("recent_folders", [])
folder_dropdown["values"] = recent_folders
favorite_folders = config.get("favorite_folders", [])
refresh_favorite_menu()
tree_font_size = config.get("font_size", 10)  # default 10
tree_font_name = "Malgun Gothic"  # or "Arial Unicode MS" for mac/Linux

tree_font = tkfont.Font(family=tree_font_name, size=tree_font_size)

style = ttk.Style()
style.configure("Treeview", 
    font=tree_font,
    rowheight=tree_font_size + 8  # Increase height based on font size
)
style.configure("Treeview.Heading", font=(tree_font_name, tree_font_size, "bold"))


result_tree.heading("#0", text="File / Sheet", anchor="w")
MIN_WIDTH = 40
for col in column_names:
    result_tree.heading(col, text=col)
    result_tree.column(col, width=100, minwidth=MIN_WIDTH, anchor="w", stretch=False)


scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=result_tree.yview)
scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=result_tree.xview)
result_tree.configure(yscroll=scroll_y.set, xscroll=scroll_x.set)

result_tree.grid(row=0, column=0, sticky="nsew")
scroll_y.grid(row=0, column=1, sticky="ns")
scroll_x.grid(row=1, column=0, sticky="ew")



tree_frame.grid_rowconfigure(0, weight=1)
tree_frame.grid_columnconfigure(0, weight=1)


# Enable drag-to-scroll
enable_drag_scroll(result_tree)

root.mainloop()
